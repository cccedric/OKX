# -*- coding: utf-8 -*-
import asyncio
from threading import Thread
import numpy as np
import openpyxl as op
import websockets
import datetime
import json

import okex.Market_api as Market
from get_balance import timestamp_datetime

flag = '1'  # 模拟盘

api_key = "2b5d8a7f-90aa-401c-af5a-9074ee40c7ca"
secret_key = "FD585FF21ED31853419E3960C840C278"
passphrase = "Casia123456!"


url = "wss://wspap.okx.com:8443/ws/v5/public?brokerId=9999"

candlesticks_bar = "1m"
candlesticks_channel = "candle"+candlesticks_bar

inst_id = "BTC-USDT"

channels = [{"channel": candlesticks_channel, 
             "instId": inst_id}]

buffer_len = 30

def update_write_candlesticks_buffer(candlesticks_buffer, res):
    if int(res[8]) == 1:
        candlesticks_buffer = np.append(candlesticks_buffer, res[4])
        if len(candlesticks_buffer) > buffer_len:
            candlesticks_buffer = np.delete(candlesticks_buffer, 0)
        # print("[candlesticks_buffer updated]", candlesticks_buffer)
    return candlesticks_buffer

async def main_loop(url, channels, candlesticks_buffer):
    while True:
        try:
            async with websockets.connect(url) as ws:
                sub_param = {"op": "subscribe", "args": channels}
                sub_str = json.dumps(sub_param)
                await ws.send(sub_str)
                print(f"send: {sub_str}")

                while True:
                    try:
                        res = await asyncio.wait_for(ws.recv(), timeout=25)
                    except (asyncio.TimeoutError, websockets.exceptions.ConnectionClosed) as e:
                        print("等待信息超时")
                        try:
                            await ws.send('ping')
                            res = await ws.recv()
                            print("[ping res]", res)
                            continue
                        except Exception as e:
                            print("连接关闭，正在重连……")
                            break

                    res = eval(res)
                    # print(res)
                    if 'event' in res:
                        continue
                    candlesticks_buffer = update_write_candlesticks_buffer(candlesticks_buffer, res['data'][0])

        except Exception as e:
            print("[Error]", e)
            print("连接断开，正在重连……")
            continue

if __name__ == '__main__':
    # marketAPI = Market.MarketAPI(api_key, secret_key, passphrase, False, flag)
    # result = marketAPI.get_history_candlesticks(instId=inst_id, limit=buffer_len)
    # print(result)
    # wb = op.load_workbook("test.xlsx")
    # ws = wb['Sheet1']
    # print(len(result['data']))
    # for re in result['data']:
    #     ws.insert_rows(0)
    #     ws['A1'] = str(timestamp_datetime(int(re[0])))
    #     ws['B1'] = float(re[1])
    #     ws['C1'] = float(re[2])
    #     ws['D1'] = float(re[3])
    #     ws['E1'] = float(re[4])
    #     ws['F1'] = float(re[5])
    #     ws['G1'] = float(re[6])
    #     ws['H1'] = float(re[7])
    #     ws['I1'] = float(re[8])

    # wb.save("test.xlsx")

    candlesticks_buffer = np.zeros([0])

    marketAPI = Market.MarketAPI(api_key, secret_key, passphrase, False, flag)
    candle_history = marketAPI.get_history_candlesticks(instId=inst_id, limit=buffer_len, bar=candlesticks_bar)
    for re in candle_history['data']:
        candlesticks_buffer = update_write_candlesticks_buffer(candlesticks_buffer, re)

    loop = asyncio.get_event_loop()
    
    loop.run_until_complete(main_loop(url, channels, candlesticks_buffer))

    loop.close()




